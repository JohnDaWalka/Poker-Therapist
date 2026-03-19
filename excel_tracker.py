"""
Excel Poker Tracker
===================
Import and export poker session data between Excel (.xlsx) workbooks
and the PlayerStatsCache SQLite database.

Expected workbook layout (sheet: "Sessions"):

    | Date       | Player ID   | VPIP | PFR | 3Bet | AF  | WTSD | Hands |
    |------------|-------------|------|-----|------|-----|------|-------|
    | 2024-01-15 | Villain_87  |  31  |  24 |    9 | 2.5 |   28 |   842 |

A second sheet "Summary" is written (or refreshed) automatically when
you call :meth:`ExcelPokerTracker.export_to_excel`.

Usage – import from Excel:
    from excel_tracker import ExcelPokerTracker
    from player_stats_cache import PlayerStatsCache

    tracker = ExcelPokerTracker("my_tracker.xlsx")
    cache   = PlayerStatsCache()
    tracker.import_to_cache(cache)

Usage – export cache to Excel:
    tracker.export_from_cache(cache, "output_tracker.xlsx")

Usage – create a blank template:
    ExcelPokerTracker.create_template("tracker_template.xlsx")
"""

from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Any, List, Optional

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover
    _OPENPYXL_AVAILABLE = False

try:
    from .player_stats_cache import PlayerStatsCache
except ImportError:
    from player_stats_cache import PlayerStatsCache

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

_SESSION_COLUMNS = [
    "Date", "Player ID", "VPIP", "PFR", "3Bet", "AF", "WTSD", "Hands",
]

_HEADER_FILL   = PatternFill("solid", fgColor="1F4E79") if _OPENPYXL_AVAILABLE else None
_HEADER_FONT   = Font(bold=True, color="FFFFFF")        if _OPENPYXL_AVAILABLE else None
_SUMMARY_FILL  = PatternFill("solid", fgColor="2E75B6") if _OPENPYXL_AVAILABLE else None


# ---------------------------------------------------------------------------
# Public class
# ---------------------------------------------------------------------------

class ExcelPokerTracker:
    """Read and write poker tracker data in Excel (.xlsx) format.

    Args:
        path: Path to the ``.xlsx`` workbook.  Need not exist yet.
    """

    SESSIONS_SHEET  = "Sessions"
    SUMMARY_SHEET   = "Summary"

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)

    # ------------------------------------------------------------------
    # Import
    # ------------------------------------------------------------------

    def read_sessions(self) -> List[Dict[str, Any]]:
        """Parse the *Sessions* sheet and return rows as a list of dicts.

        Returns:
            Each dict has keys: ``player_id``, ``vpip``, ``pfr``,
            ``threebet``, ``af``, ``wtsd``, ``hands``, and optionally
            ``date``.

        Raises:
            FileNotFoundError: If the workbook does not exist.
            ValueError:        If the *Sessions* sheet or expected header
                               columns are absent.
        """
        _require_openpyxl()
        if not self.path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.path}")

        wb = load_workbook(self.path, read_only=True, data_only=True)
        if self.SESSIONS_SHEET not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{self.SESSIONS_SHEET}' not found in {self.path}. "
                f"Available sheets: {wb.sheetnames}"
            )

        ws = wb[self.SESSIONS_SHEET]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        # Build a case-insensitive column index from the header row
        header = [str(h).strip() if h is not None else "" for h in rows[0]]
        col_idx = {_normalise_col(h): i for i, h in enumerate(header)}

        required = {"player id", "vpip", "pfr"}
        missing  = required - col_idx.keys()
        if missing:
            raise ValueError(
                f"Missing required columns in '{self.SESSIONS_SHEET}': "
                f"{', '.join(sorted(missing))}"
            )

        sessions: List[Dict[str, Any]] = []
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue  # skip blank rows

            def _get(name: str, default=None):
                idx = col_idx.get(name)
                return row[idx] if idx is not None and idx < len(row) else default

            player_id = str(_get("player id", "")).strip()
            if not player_id:
                continue

            sessions.append({
                "player_id": player_id,
                "vpip":      _to_float(_get("vpip",  0)),
                "pfr":       _to_float(_get("pfr",   0)),
                "threebet":  _to_float(_get("3bet",  0)),
                "af":        _to_float(_get("af",    0)),
                "wtsd":      _to_float(_get("wtsd",  0)),
                "hands":     _to_int(_get("hands",   0)),
                "date":      _get("date"),
            })

        wb.close()
        return sessions

    def import_to_cache(self, cache: PlayerStatsCache) -> int:
        """Import all sessions from the workbook into *cache*.

        For players that appear multiple times the stats are aggregated
        (hands summed; VPIP/PFR/3Bet/AF/WTSD averaged weighted by hands)
        before a single upsert is sent to the cache.

        Args:
            cache: A :class:`PlayerStatsCache` instance to write into.

        Returns:
            Number of distinct players upserted.
        """
        sessions = self.read_sessions()
        aggregated = _aggregate_sessions(sessions)

        for player_id, stats in aggregated.items():
            cache.update_stats(
                player_id,
                vpip=stats["vpip"],
                pfr=stats["pfr"],
                threebet=stats["threebet"],
                af=stats["af"],
                wtsd=stats["wtsd"],
                hands=stats["hands"],
            )

        return len(aggregated)

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    def export_from_cache(
        self,
        cache: PlayerStatsCache,
        output_path: Optional[str | Path] = None,
    ) -> Path:
        """Write all players from *cache* into an Excel workbook.

        Creates (or overwrites) the workbook at *output_path* (defaults to
        :attr:`path`).

        Args:
            cache:       Source :class:`PlayerStatsCache`.
            output_path: Destination ``.xlsx`` file (optional).

        Returns:
            Path to the written workbook.
        """
        _require_openpyxl()
        dest = Path(output_path) if output_path else self.path
        players = cache.get_all_players()

        wb = Workbook()
        _write_sessions_sheet(wb, players)
        _write_summary_sheet(wb, players)

        dest.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(dest))
        return dest

    # ------------------------------------------------------------------
    # Template
    # ------------------------------------------------------------------

    @staticmethod
    def create_template(output_path: str | Path = "tracker_template.xlsx") -> Path:
        """Create a blank template workbook that shows the expected layout.

        Args:
            output_path: Where to write the template.

        Returns:
            Path to the written template.
        """
        _require_openpyxl()
        dest = Path(output_path)
        wb   = Workbook()

        # --- Sessions sheet ---
        ws = wb.active
        ws.title = ExcelPokerTracker.SESSIONS_SHEET

        _apply_header_row(ws, _SESSION_COLUMNS)

        # Example rows
        examples = [
            ["2024-01-15", "Villain_87",  31, 24,  9, 2.5, 28, 842],
            ["2024-01-16", "NitPlayer",   12,  9,  2, 1.2, 22, 315],
            ["2024-01-17", "ManiacMike",  48, 38, 18, 4.1, 32, 190],
        ]
        for row_data in examples:
            ws.append(row_data)

        _autofit_columns(ws)

        # --- Summary placeholder ---
        ws_sum = wb.create_sheet(ExcelPokerTracker.SUMMARY_SHEET)
        ws_sum["A1"] = "Run 'export_from_cache()' to populate this sheet."
        ws_sum["A1"].font = Font(italic=True, color="808080")

        dest.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(dest))
        return dest


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _require_openpyxl() -> None:
    if not _OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel support. "
            "Install it with: pip install openpyxl"
        )


def _normalise_col(name: str) -> str:
    """Lowercase and collapse whitespace for fuzzy column matching."""
    return re.sub(r"\s+", " ", name.strip().lower())


def _to_float(value, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _to_int(value, default: int = 0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _aggregate_sessions(
    sessions: List[Dict[str, Any]],
) -> Dict[str, Dict[str, Any]]:
    """Aggregate multiple session rows for the same player.

    Stats are weighted by hands count so that a 1,000-hand session
    counts more than a 10-hand sample.
    """
    buckets: Dict[str, Dict[str, Any]] = {}

    for s in sessions:
        pid   = s["player_id"]
        hands = s["hands"]
        if pid not in buckets:
            buckets[pid] = {
                "vpip": 0.0, "pfr": 0.0, "threebet": 0.0,
                "af":   0.0, "wtsd": 0.0, "hands": 0,
                "_weight": 0.0,
            }
        b = buckets[pid]
        w = max(hands, 1)  # avoid division by zero when calculating weighted averages
        for stat in ("vpip", "pfr", "threebet", "af", "wtsd"):
            b[stat] = (b[stat] * b["_weight"] + s[stat] * w) / (b["_weight"] + w)
        b["hands"]   += hands
        b["_weight"] += w

    # Drop the internal weight key
    for b in buckets.values():
        del b["_weight"]

    return buckets


def _apply_header_row(ws, columns: List[str]) -> None:
    ws.append(columns)
    for col_num, _ in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.font  = _HEADER_FONT
        cell.fill  = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _autofit_columns(ws, min_width: int = 10) -> None:
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col),
            default=min_width,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(
            max_len + 2, min_width
        )


def _write_sessions_sheet(wb: "Workbook", players: List[Dict[str, Any]]) -> None:
    ws = wb.active
    ws.title = ExcelPokerTracker.SESSIONS_SHEET

    _apply_header_row(ws, _SESSION_COLUMNS)

    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for p in players:
        ws.append([
            now_str,
            p.get("player_id", ""),
            round(p.get("vpip", 0), 1),
            round(p.get("pfr",  0), 1),
            round(p.get("threebet", 0), 1),
            round(p.get("af",   0), 2),
            round(p.get("wtsd", 0), 1),
            p.get("hands", 0),
        ])

    _autofit_columns(ws)


def _write_summary_sheet(wb: "Workbook", players: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(ExcelPokerTracker.SUMMARY_SHEET)

    summary_cols = ["Player ID", "Style", "VPIP", "PFR", "3Bet", "Hands"]
    _apply_header_row(ws, summary_cols)
    ws["A1"].fill = ws["B1"].fill = ws["C1"].fill = _SUMMARY_FILL
    ws["D1"].fill = ws["E1"].fill = ws["F1"].fill = _SUMMARY_FILL

    for p in players:
        ws.append([
            p.get("player_id",    ""),
            p.get("player_style", "Unknown"),
            round(p.get("vpip",     0), 1),
            round(p.get("pfr",      0), 1),
            round(p.get("threebet", 0), 1),
            p.get("hands", 0),
        ])

    _autofit_columns(ws)


# ---------------------------------------------------------------------------
# CLI / self-test
# ---------------------------------------------------------------------------

def _run_tests() -> None:
    """Run built-in Excel tracker tests."""
    import tempfile
    import shutil
    from pathlib import Path

    print("Running Excel tracker tests...")

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)

        # 1. create_template ---------------------------------------------------
        tmpl = ExcelPokerTracker.create_template(tmp_path / "template.xlsx")
        assert tmpl.exists(), "Template file not created"
        wb = load_workbook(tmpl)
        assert ExcelPokerTracker.SESSIONS_SHEET in wb.sheetnames, \
            "Sessions sheet missing from template"
        print("  ✓ create_template writes workbook with Sessions sheet")

        # 2. read_sessions from template ---------------------------------------
        tracker = ExcelPokerTracker(tmpl)
        sessions = tracker.read_sessions()
        assert len(sessions) == 3, f"Expected 3 example rows, got {len(sessions)}"
        assert sessions[0]["player_id"] == "Villain_87", "First player mismatch"
        assert sessions[0]["vpip"] == 31.0, "VPIP mismatch"
        assert sessions[0]["hands"] == 842, "Hands mismatch"
        print("  ✓ read_sessions parses example rows correctly")

        # 3. import_to_cache ---------------------------------------------------
        db_path = tmp_path / "test.db"
        cache   = PlayerStatsCache(db_path=db_path)
        count   = tracker.import_to_cache(cache)
        assert count == 3, f"Expected 3 players imported, got {count}"
        villain = cache.get_player_stats("Villain_87")
        assert villain is not None, "Villain_87 not found in cache"
        assert villain["hands"] == 842, "Hands not imported correctly"
        print("  ✓ import_to_cache upserts all players")

        # 4. aggregation (same player in multiple rows) ------------------------
        multi_path = tmp_path / "multi.xlsx"
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Sessions"
        ws2.append(_SESSION_COLUMNS)
        ws2.append(["2024-01-01", "TestPlayer", 20, 15, 5, 1.5, 25, 100])
        ws2.append(["2024-01-02", "TestPlayer", 30, 20, 8, 2.0, 30, 300])
        wb2.save(str(multi_path))

        tracker2  = ExcelPokerTracker(multi_path)
        sessions2 = tracker2.read_sessions()
        agg = _aggregate_sessions(sessions2)
        assert agg["TestPlayer"]["hands"] == 400, "Hands not summed correctly"
        # Weighted avg VPIP: (20*100 + 30*300) / 400 = (2000+9000)/400 = 27.5
        assert abs(agg["TestPlayer"]["vpip"] - 27.5) < 0.01, \
            f"Weighted VPIP incorrect: {agg['TestPlayer']['vpip']}"
        print("  ✓ aggregate_sessions sums hands and weights averages")

        # 5. export_from_cache -------------------------------------------------
        out_path = tmp_path / "export.xlsx"
        tracker3 = ExcelPokerTracker(out_path)
        written  = tracker3.export_from_cache(cache, out_path)
        assert written.exists(), "Exported workbook not created"
        wb3 = load_workbook(written)
        assert ExcelPokerTracker.SESSIONS_SHEET in wb3.sheetnames, \
            "Sessions sheet missing from export"
        assert ExcelPokerTracker.SUMMARY_SHEET in wb3.sheetnames, \
            "Summary sheet missing from export"
        print("  ✓ export_from_cache writes Sessions and Summary sheets")

        # 6. FileNotFoundError on missing workbook ----------------------------
        try:
            ExcelPokerTracker(tmp_path / "ghost.xlsx").read_sessions()
            assert False, "Expected FileNotFoundError"
        except FileNotFoundError:
            pass
        print("  ✓ read_sessions raises FileNotFoundError for missing file")

        # 7. ValueError on missing Sessions sheet -----------------------------
        bad_path = tmp_path / "bad.xlsx"
        wb_bad = Workbook()
        wb_bad.active.title = "WrongSheet"
        wb_bad.save(str(bad_path))
        try:
            ExcelPokerTracker(bad_path).read_sessions()
            assert False, "Expected ValueError"
        except ValueError:
            pass
        print("  ✓ read_sessions raises ValueError for missing Sessions sheet")

    print("\n✅ All Excel tracker tests passed!")


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Excel Poker Tracker")
        print("=" * 30)
        print("\nUsage:")
        print("  python excel_tracker.py template [output.xlsx]")
        print("  python excel_tracker.py import   <tracker.xlsx> [db.sqlite]")
        print("  python excel_tracker.py export   <db.sqlite>    [output.xlsx]")
        print("  python excel_tracker.py test")
        sys.exit(0)

    cmd = sys.argv[1]

    if cmd == "template":
        dest = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("tracker_template.xlsx")
        ExcelPokerTracker.create_template(dest)
        print(f"✓ Template created: {dest}")

    elif cmd == "import":
        if len(sys.argv) < 3:
            print("Usage: python excel_tracker.py import <tracker.xlsx> [db.sqlite]")
            sys.exit(1)
        xlsx_path = Path(sys.argv[2])
        db_path   = Path(sys.argv[3]) if len(sys.argv) > 3 else None
        cache     = PlayerStatsCache(db_path=db_path)
        tracker   = ExcelPokerTracker(xlsx_path)
        n         = tracker.import_to_cache(cache)
        print(f"✓ Imported {n} player(s) from {xlsx_path}")

    elif cmd == "export":
        if len(sys.argv) < 3:
            print("Usage: python excel_tracker.py export <db.sqlite> [output.xlsx]")
            sys.exit(1)
        db_path   = Path(sys.argv[2])
        out_path  = Path(sys.argv[3]) if len(sys.argv) > 3 else Path("export_tracker.xlsx")
        cache     = PlayerStatsCache(db_path=db_path)
        tracker   = ExcelPokerTracker(out_path)
        written   = tracker.export_from_cache(cache, out_path)
        print(f"✓ Exported to {written}")

    elif cmd == "test":
        _run_tests()
        sys.exit(0)

    else:
        print(f"Unknown command: {cmd}")
        sys.exit(1)
